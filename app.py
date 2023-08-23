from flask import Flask, render_template, request, send_from_directory
import os
import pandas as pd
from openpyxl import load_workbook

# Initialize the Flask application
app = Flask(__name__)

# This is the path to the upload directory,
# which is temporary and will be deleted afterwards
# create tmp folder if it doesn't exist
UPLOAD_FOLDER = "tmp"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Set the folder to save the tmp file to
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Initialize the templates directory
app.template_folder = "templates"


# This will be called after each request to delete the tmp directory
def clean_tmp():
    for filename in os.listdir(app.config["UPLOAD_FOLDER"]):
        os.remove(os.path.join(app.config["UPLOAD_FOLDER"], filename))


def extract_flagged_sales_orders(file_path, output_path, group_column="Sales Order"):
    """
    Extracts and formats data from an Excel file based on the selected grouping column.

    Args:
    - file_path (str): Path to the input Excel file.
    - output_path (str): Path where the processed file will be saved.
    - group_column (str): The column by which the data should be grouped.
    """

    # Load the workbook and active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Extract highlighted cell coordinates
    flagged_sales_orders = []
    row_index = 1  # Starting from 1 to match Excel's 1-based indexing

    # Blue FF00B0F0 # green FF92D050 # red FFFFC7CE
    for row in ws.iter_rows():
        if row[4].fill.start_color.rgb == "FFFFFFFF":  # if the cell is white
            row_index += 1  # increment the row index
            print(row_index, row[4].value)
            continue  # skip the row

        else:
            if (
                row[0].fill.start_color.rgb == "FFFFC7CE"
            ):  # stop when we reach the light rec color
                continue  # stop iterating through the rows when we reach the light red color

            if row[3].fill.start_color.rgb == "FF92D050":  # if the cell is green
                # Adjust the index by subtracting 1 to account for zero-based indexing in pandas
                flagged_sales_orders.append(row_index - 1)

        row_index += 1  # increment the row index

    # Load the data into a DataFrame
    df = pd.read_excel(file_path)

    # Filter DataFrame to only include the columns of interest
    columns_of_interest = [
        "Sales Order",
        "Item ID",
        "Required Qty",
    ]  # HERE is where to modify the columns in the output

    # filter the dataframe to only include the columns of interest
    df = df[columns_of_interest]

    # Filter DataFrame to only include the rows of interest
    filtered_df = df[df.index.isin(flagged_sales_orders)]

    # Sort by 'Sales Order' in ascending order
    sorted_df = filtered_df.sort_values(by="Sales Order", ascending=True).reset_index(
        drop=True
    )

    # Add empty rows between groups
    grouped = (
        sorted_df.groupby(group_column, group_keys=True)
        .apply(lambda x: x._append(pd.Series(), ignore_index=True))
        .reset_index(drop=True)
    )

    # Save to new Excel file
    grouped.to_excel(output_path, index=False)


# Main route
@app.route("/", methods=["GET", "POST"])
def index():
    """
    Flask endpoint for the main page.
    Renders the upload page on GET request.
    Processes the uploaded file and returns the processed file as a download on POST request.
    """
    if request.method == "POST":
        file = request.files["file"]
        selected_column = request.form.get("groupColumn")
        if file:
            filename = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
            file.save(filename)

            output_path = os.path.join(
                app.config["UPLOAD_FOLDER"], "processed_output.xlsx"
            )
            extract_flagged_sales_orders(filename, output_path, selected_column)

            return send_from_directory(
                app.config["UPLOAD_FOLDER"], "processed_output.xlsx", as_attachment=True
            )

    # If GET request, render the upload page
    clean_tmp()
    return render_template("uploads.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=9876)
