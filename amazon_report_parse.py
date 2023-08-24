import sys
from openpyxl import load_workbook, Workbook


def create_workbook(file_path):
    """
    Load and return a workbook from the given file path.

    Args:
    - file_path (str): Path to the Excel file.

    Returns:
    - openpyxl.worksheet.worksheet.Worksheet: Loaded Excel worksheet.
    """
    wb = load_workbook(file_path)
    return wb.active


def filter_rows_by_fill_color(ws, color):
    """
    Filter and return rows based on the fill color of their cells in column D.

    Args:
    - ws (openpyxl.worksheet.worksheet.Worksheet): Loaded Excel worksheet.
    - color (str): Fill color to filter by.

    Returns:
    - list: List of rows with cells in column D having the specified fill color.
    """
    selected_rows = []
    for row in ws.iter_rows():
        if row[3].fill.fgColor.rgb == color:  # 3 corresponds to column D
            selected_rows.append(row)
    return selected_rows


def save_filtered_data_to_excel(sorted_rows, output_file):
    """
    Save the filtered and sorted rows to a new Excel file with specific columns.

    Args:
    - sorted_rows (list): Sorted rows to save.
    - output_file (str): Path to the output Excel file.
    """
    # Create a new workbook and worksheet
    new_wb = Workbook()
    new_ws = new_wb.active

    # Add headers
    new_ws.append(["Sales Order", "Item ID", "Required Qty"])

    # Add the filtered and sorted data
    for row in sorted_rows:
        if row != ["", "", ""]:
            new_ws.append([row[3].value, row[6].value, row[13].value])
        else:
            new_ws.append(row)

    # Save the new workbook
    new_wb.save(output_file)


def sort_by_column_az(rows, asc=True):
    """
    Sort the rows by the values in column D in ascending order.

    Args:
    - rows (list): List of rows to sort.
    - asc (bool): Whether to sort in ascending order (True) or descending order (False).

    Returns:
    - list: Sorted rows.
    """
    return sorted(rows, key=lambda x: x[3].value, reverse=not asc)


def separate_groups(rows):
    """
    Separate the rows into groups based on the values in column D.

    Args:
    - rows (list): List of rows to separate.

    Returns:
    - list: List of rows separated into groups.
    """
    parsed_rows = []
    for group in group_rows(rows):
        parsed_rows.extend(group)
        parsed_rows.append(["", "", ""])  # add a blank row between groups
    return parsed_rows


def group_rows(rows):
    """
    Group the rows based on the values in column D.

    Args:
    - rows (list): List of rows to group.

    Returns:
    - list: List of rows grouped by the values in column D.
    """
    last_row = None
    group = []
    for row in rows:
        if last_row is None:
            group.append(row)
        elif row[3].value == last_row[3].value:
            group.append(row)
        else:
            yield group
            group = [row]
        last_row = row
    yield group


def parse_amazon_report(filepath=None, output_file=None):
    """
    Parse the Amazon report and save the parsed data to an Excel file.

    Args:
    - filepath (str): Path to the Amazon report Excel file.
    - output_file (str): Path to the output Excel file.
    """
    if filepath is None:
        file_path = sys.argv[1]
    if output_file is None:
        output_file = "parsed_output.xlsx"  # Default output file name
        if len(sys.argv) > 2:
            output_file = sys.argv[2]

    # create a workbook and worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # extract highlighted cell coordinates
    color = "FF92D050"  # green color code: FF92D050
    filtered = filter_rows_by_fill_color(ws, color)
    sorted_rows = sort_by_column_az(filtered)
    grouped_and_sorted_rows = separate_groups(sorted_rows)
    save_filtered_data_to_excel(grouped_and_sorted_rows, output_file)


if __name__ == "__main__":
    parse_amazon_report()
