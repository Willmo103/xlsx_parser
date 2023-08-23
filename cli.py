import sys
from openpyxl import load_workbook
from app import extract_flagged_sales_orders

# filter all rows by the color of column D
# and print each value in that column


def create_workbook(file_path):
    """
    Load and return a workbook from the given file path.

    Args:
    - file_path (str): Path to the Excel file.

    Returns:
    - openpyxl.worksheet.worksheet.Worksheet: Loaded Excel worksheet.
    """
    wb = load_workbook(file_path)
    return wb.active


def filter_rows_by_fill_color(file_path, color):
    """
    Filter and return rows based on the fill color of their cells.

    Args:
    - file_path (str): Path to the Excel file.
    - color (str): Fill color to filter by.

    Returns:
    - list: List of rows with cells having the specified fill color.
    """
    selected_rows = []
    ws = create_workbook(file_path)
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.rgb == color:
                selected_rows.append(row)
                break
    return selected_rows


def sort_rows_by_value(selected_rows, column_index):
    """
    Sort the rows based on the values of a specific column.

    Args:
    - selected_rows (list): List of rows to sort.
    - column_index (int): Index of the column to sort by.

    Returns:
    - list: Sorted list of rows.
    """
    return sorted(selected_rows, key=lambda x: x[column_index].value)


if __name__ == "__main__":
    ws_path = sys.argv[1]
    color = "FF92D050"  # green color code: FF92D050
    filtered_rows = filter_rows_by_fill_color(ws_path, color)
    sorted_rows = sort_rows_by_value(
        filtered_rows, 3
    )  # Index 3 corresponds to column D
    for row in sorted_rows:
        # Print the value of the cell in column D for each row
        print(row[3].value)
